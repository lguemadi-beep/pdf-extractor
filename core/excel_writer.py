"""
excel_writer.py
----------------
Takes a list of PdfResult objects (see extractor.py) and writes a
professionally formatted .xlsx workbook:

  - "Summary"     : one row per PDF (name, pages, #tables, #fields, status)
  - "Fields"      : every Key/Value pair found, one row per pair
  - "Tables"      : every table found, stacked with source file/page columns
  - "Table_<n>"   : (optional) each table also gets its own clean sheet

Formatting: header styling, frozen header row, autosized columns,
banded rows, autofilter, and a document-wide color theme.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .extractor import PdfResult
from .facture_parser import Facture

# ---------------------------------------------------------------------------
# Theme
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
TITLE_FONT = Font(color="1F4E78", bold=True, size=16, name="Calibri")
SUBTITLE_FONT = Font(color="595959", italic=True, size=10, name="Calibri")
BAND_FILL = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
TOTAL_FONT = Font(bold=True, size=11, name="Calibri")
TOTAL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
NUMBER_FORMAT = "#,##0.00"


def _write_header(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def _band_rows(ws: Worksheet, start_row: int, end_row: int, n_cols: int) -> None:
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = BAND_FILL
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if ws.cell(row=r, column=c).alignment.horizontal is None:
                ws.cell(row=r, column=c).alignment = LEFT


def _title_block(ws: Worksheet, title: str, subtitle: str, span_cols: int) -> int:
    """Write a title + subtitle at the top of a sheet, return next free row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT
    ws.row_dimensions[1].height = 24
    return 4  # header row starts here


def _fill_factures_sheet(ws: Worksheet, factures: list[Facture], source_folder: str) -> Worksheet:
    total_lignes = sum(len(f.lignes) for f in factures)
    header_row = _title_block(
        ws,
        "Factures Proforma",
        f"Dossier source : {source_folder}  |  Factures : {len(factures)}  |  Lignes : {total_lignes}",
        span_cols=6,
    )
    headers = ["Fournisseur", "Demandeur", "Numéro facture", "Date de facture", "Description", "Montant HT"]
    _write_header(ws, headers, row=header_row)
    r = header_row + 1
    for fac in factures:
        if fac.lignes:
            for ligne in fac.lignes:
                row_vals = [fac.societe, fac.demandeur, fac.numero, fac.date, ligne.designation, ligne.montant_ht]
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r, column=c_idx, value=val)
                    if c_idx == 6 and isinstance(val, (int, float)):
                        cell.number_format = NUMBER_FORMAT
                r += 1
        else:
            # Invoice with no parsed line items: still show one row so it isn't lost
            row_vals = [fac.societe, fac.demandeur, fac.numero, fac.date, "", fac.sous_total_ht]
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                if c_idx == 6 and isinstance(val, (int, float)):
                    cell.number_format = NUMBER_FORMAT
            r += 1

    # Grand-total row
    if factures:
        total_row = r
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        total_ht = sum(l.montant_ht for f in factures for l in f.lignes) or sum(f.sous_total_ht or 0 for f in factures)
        cell = ws.cell(row=total_row, column=6, value=total_ht)
        cell.number_format = NUMBER_FORMAT
        cell.font = TOTAL_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=c).fill = TOTAL_FILL
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        r += 1

    _band_rows(ws, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws)
    return ws


def write_workbook(
    results: list[PdfResult],
    output_path: Path,
    source_folder: str,
    factures: list[Facture] | None = None,
) -> Path:
    factures = factures or []
    matched_names = {f.file_name for f in factures}
    generic_results = [r for r in results if r.file_name not in matched_names]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # we'll add sheets explicitly in the right order

    # ------------------------------------------------------- Factures / Lignes
    if factures:
        ws_fac = wb.create_sheet("Factures")
        _fill_factures_sheet(ws_fac, factures, source_folder)
        wb.active = 0

    # If nothing was auto-recognized, fall back fully to the generic report
    if not generic_results:
        if not factures:
            wb.create_sheet("Summary")  # keep a sheet so the workbook isn't empty
        wb.save(str(output_path))
        return output_path

    # ---------------------------------------------------------------- Summary
    ws_summary = wb.create_sheet("Summary")
    header_row = _title_block(
        ws_summary,
        "PDF Extraction Summary",
        f"Source folder: {source_folder}  |  Files (not matching a known template): {len(generic_results)}  |  "
        f"Generated: {generic_results[0].extracted_at if generic_results else ''}",
        span_cols=7,
    )
    headers = ["File name", "Pages", "Tables found", "Fields found", "Dates found", "Amounts found", "Status"]
    _write_header(ws_summary, headers, row=header_row)
    r = header_row + 1
    for res in generic_results:
        status = "Error" if res.error else "OK"
        row_vals = [
            res.file_name,
            res.page_count,
            len(res.tables),
            len(res.key_values),
            len(res.dates_found),
            len(res.amounts_found),
            res.error if res.error else status,
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_summary.cell(row=r, column=c_idx, value=val)
            if res.error:
                cell.fill = ERROR_FILL
        r += 1
    _band_rows(ws_summary, header_row + 1, r - 1, len(headers))
    _autosize(ws_summary)

    # ----------------------------------------------------------------- Fields
    ws_fields = wb.create_sheet("Fields")
    header_row = _title_block(
        ws_fields, "Extracted Key / Value Fields", "One row per field found in each document.", span_cols=4
    )
    headers = ["File name", "Page count", "Field", "Value"]
    _write_header(ws_fields, headers, row=header_row)
    r = header_row + 1
    for res in generic_results:
        for key, value in res.key_values.items():
            ws_fields.cell(row=r, column=1, value=res.file_name)
            ws_fields.cell(row=r, column=2, value=res.page_count)
            ws_fields.cell(row=r, column=3, value=key)
            ws_fields.cell(row=r, column=4, value=value)
            r += 1
    _band_rows(ws_fields, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws_fields)

    # ----------------------------------------------------------------- Tables
    ws_tables = wb.create_sheet("Tables")
    header_row = _title_block(
        ws_tables, "Extracted Tables (combined)", "All tables from all documents, stacked together.", span_cols=6
    )
    headers = ["File name", "Page", "Table #", "Row #", "Column", "Value"]
    _write_header(ws_tables, headers, row=header_row)
    r = header_row + 1
    for res in generic_results:
        for tbl in res.tables:
            for row_idx, row in enumerate(tbl.rows, start=1):
                for col_idx, val in enumerate(row):
                    col_name = tbl.headers[col_idx] if col_idx < len(tbl.headers) else f"col_{col_idx+1}"
                    ws_tables.cell(row=r, column=1, value=res.file_name)
                    ws_tables.cell(row=r, column=2, value=tbl.page_number)
                    ws_tables.cell(row=r, column=3, value=tbl.table_index)
                    ws_tables.cell(row=r, column=4, value=row_idx)
                    ws_tables.cell(row=r, column=5, value=col_name)
                    ws_tables.cell(row=r, column=6, value=val)
                    r += 1
    _band_rows(ws_tables, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws_tables)

    # ------------------------------------------------- One clean sheet / table
    used_names = {"Factures", "Summary", "Fields", "Tables"}
    for res in generic_results:
        for tbl in res.tables:
            base = f"{Path(res.file_name).stem}_p{tbl.page_number}_t{tbl.table_index}"
            sheet_name = base[:31]
            suffix = 1
            while sheet_name in used_names:
                sheet_name = f"{base[:28]}_{suffix}"[:31]
                suffix += 1
            used_names.add(sheet_name)

            ws_t = wb.create_sheet(sheet_name)
            header_row_t = _title_block(
                ws_t, f"{res.file_name}", f"Page {tbl.page_number} — Table {tbl.table_index}", span_cols=max(1, len(tbl.headers))
            )
            _write_header(ws_t, tbl.headers, row=header_row_t)
            rr = header_row_t + 1
            for row in tbl.rows:
                for c_idx, val in enumerate(row, start=1):
                    ws_t.cell(row=rr, column=c_idx, value=val)
                rr += 1
            _band_rows(ws_t, header_row_t + 1, max(rr - 1, header_row_t + 1), max(1, len(tbl.headers)))
            _autosize(ws_t)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
