"""
excel_writer.py
----------------
Takes a list of PdfResult objects (see extractor.py) and writes a
professionally formatted .xlsx workbook:

  - "Summary"     : one row per PDF (name, pages, #tables, #fields, status)
  - "Fields"      : every Key/Value pair found, one row per pair
  - "Tables"      : every table found, stacked with source file/page columns
  - "Table_<n>"   : (optional) each table also gets its own clean sheet

Formatting: header styling, frozen header row, autosized columns,
banded rows, autofilter, and a document-wide color theme.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .extractor import PdfResult
from .facture_parser import Facture

logger = logging.getLogger("pdf_extractor")

# ---------------------------------------------------------------------------
# Theme
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
TITLE_FONT = Font(color="1F4E78", bold=True, size=16, name="Calibri")
SUBTITLE_FONT = Font(color="595959", italic=True, size=10, name="Calibri")
BAND_FILL = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
TOTAL_FONT = Font(bold=True, size=11, name="Calibri")
TOTAL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
NUMBER_FORMAT = "#,##0.00"

# Fixed dropdown choices for the "Owner CPM / BA" column.
OWNER_OPTIONS = ["Racha Chebbour", "Med Hacene Riadh", "Kenza Ghidouche", "Salima Saidi"]

# Fixed dropdown choices for the "Status" column.
STATUS_OPTIONS = ["Transmis pour la DA", "Basculer vers le PO", "Faire le GR", "GR Done", "Déposer la facture"]


def _add_dropdown(ws: Worksheet, options: list[str], column_index: int, first_row: int, last_row: int = 1000, title: str = "") -> None:
    """Restrict a column to a fixed list of options via an Excel dropdown."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
    )
    dv.error = "Choisissez une valeur dans la liste."
    dv.errorTitle = "Valeur invalide"
    dv.prompt = "Sélectionnez une valeur dans la liste."
    dv.promptTitle = title
    col_letter = get_column_letter(column_index)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    ws.add_data_validation(dv)


def _protect_extracted_columns(
    ws: Worksheet,
    locked_columns: list[int],
    editable_columns: list[int],
    first_row: int,
    last_row: int = 1000,
) -> None:
    """
    Turns on sheet protection so the extracted-data columns are read-only,
    while the manual tracking columns (Owner CPM / BA, Affectation, Status)
    stay editable. No password is set — this guards against accidental
    edits, not a security measure (Review > Unprotect Sheet removes it).
    """
    for row in range(first_row, last_row + 1):
        for col in locked_columns:
            ws.cell(row=row, column=col).protection = Protection(locked=True)
        for col in editable_columns:
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.autoFilter = False  # keep the filter dropdowns usable while protected
    ws.protection.sort = False  # keep sorting usable while protected
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def _write_header(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 60, start_row: int = 1) -> None:
    """
    Auto-size each column from its actual data, ignoring rows above
    start_row — this matters because the merged title/subtitle block at
    the top of each sheet can contain a long string (e.g. "Dossier
    source: ... | Factures: 21 | Lignes: 26") that would otherwise
    inflate the first column's width far beyond its real content.
    """
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) for c in col_cells if c.value is not None and c.row >= start_row),
            default=0,
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def _band_rows(ws: Worksheet, start_row: int, end_row: int, n_cols: int) -> None:
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = BAND_FILL
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if ws.cell(row=r, column=c).alignment.horizontal is None:
                ws.cell(row=r, column=c).alignment = LEFT


def _title_block(ws: Worksheet, title: str, subtitle: str, span_cols: int) -> int:
    """Write a title + subtitle at the top of a sheet, return next free row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT
    ws.row_dimensions[1].height = 24
    return 4  # header row starts here


def _facture_to_rows(fac: Facture) -> list[tuple]:
    """
    Flatten one Facture into (fournisseur, demandeur, numero, date, description,
    montant_ht, owner, affectation, status) rows. Owner/Affectation/Status are
    not extracted from the PDF — they're free-form tracking columns left blank
    (Status is left blank by default, to be picked from the dropdown) for
    manual follow-up in Excel; existing
    values already typed in for older rows are preserved on later runs.
    """
    if fac.lignes:
        return [
            (fac.societe, fac.demandeur, fac.numero, fac.date, ligne.designation, ligne.montant_ht, "", "", "")
            for ligne in fac.lignes
        ]
    # Invoice with no parsed line items: still show one row so it isn't lost
    return [(fac.societe, fac.demandeur, fac.numero, fac.date, "", fac.sous_total_ht, "", "", "")]


def _fill_factures_sheet(ws: Worksheet, rows: list[tuple], source_folder: str, invoice_count: int) -> Worksheet:
    header_row = _title_block(
        ws,
        "Factures Proforma",
        f"Dossier source : {source_folder}  |  Factures : {invoice_count}  |  Lignes : {len(rows)}",
        span_cols=9,
    )
    headers = [
        "Fournisseur", "Demandeur", "Numéro facture", "Date de facture", "Description", "Montant HT",
        "Owner CPM / BA", "Affectation", "Status",
    ]
    _write_header(ws, headers, row=header_row)
    _add_dropdown(ws, OWNER_OPTIONS, column_index=7, first_row=header_row + 1, title="Owner CPM / BA")
    _add_dropdown(ws, STATUS_OPTIONS, column_index=9, first_row=header_row + 1, title="Status")
    r = header_row + 1
    for row_vals in rows:
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if c_idx == 6 and isinstance(val, (int, float)):
                cell.number_format = NUMBER_FORMAT
        r += 1

    last_data_row = r - 1
    if last_data_row >= header_row + 1:
        _protect_extracted_columns(
            ws, locked_columns=[1, 2, 3, 4, 5, 6], editable_columns=[7, 8, 9],
            first_row=header_row + 1, last_row=last_data_row,
        )

    # Grand-total row
    if rows:
        total_row = r
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        total_ht = sum(row[5] for row in rows if isinstance(row[5], (int, float)))
        cell = ws.cell(row=total_row, column=6, value=total_ht)
        cell.number_format = NUMBER_FORMAT
        cell.font = TOTAL_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=c).fill = TOTAL_FILL
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        r += 1

    _band_rows(ws, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws, start_row=header_row)
    return ws


def _load_existing_factures_sheet(output_path: Path) -> tuple[list[tuple], set[str]]:
    """
    If invoice_summary.xlsx already exists, read back its "Factures" sheet
    (data rows only — skipping the title block, header, and TOTAL row) so
    that data already extracted in earlier runs — including any Owner /
    Affectation / Status values typed in manually — is never lost.
    Returns (existing_rows, existing_invoice_numbers).
    """
    if not output_path.exists():
        return [], set()
    try:
        existing_wb = load_workbook(str(output_path))
    except Exception:
        logger.warning("Could not read existing %s (will start fresh).", output_path.name)
        return [], set()
    if "Factures" not in existing_wb.sheetnames:
        return [], set()

    ws = existing_wb["Factures"]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Fournisseur":
            header_row = r
            break
    if header_row is None:
        return [], set()

    rows: list[tuple] = []
    numeros: set[str] = set()
    for r in range(header_row + 1, ws.max_row + 1):
        first_cell = ws.cell(row=r, column=1).value
        if first_cell is None or first_cell == "TOTAL":
            continue
        # Read 9 columns even if this row was written by an older version of
        # the app with fewer columns (Owner/Affectation/Status will read as
        # None -> normalized to "" below).
        row_vals = tuple(ws.cell(row=r, column=c).value for c in range(1, 10))
        # normalize missing trailing cells to "" without touching numeric montant_ht (index 5)
        row_vals = tuple(
            ("" if (i >= 6 and v is None) else v) for i, v in enumerate(row_vals)
        )
        rows.append(row_vals)
        numero = row_vals[2]
        if numero:
            numeros.add(str(numero))
    return rows, numeros


def _group_rows_by_invoice(rows: list[tuple]) -> list[tuple]:
    """
    Collapse line-item rows (one row per prestation) into one row per unique
    invoice number: (fournisseur, demandeur, numero, date, montant_ht_total,
    owner, affectation, status). Montant HT is summed across that invoice's
    lines; Owner/Affectation/Status take the first non-empty value found
    among its lines (manual tracking is usually set once per invoice).
    Order follows first appearance of each invoice number.
    """
    groups: dict[str, dict] = {}
    order: list[str] = []
    for row in rows:
        fournisseur, demandeur, numero, date, _description, montant_ht = row[0], row[1], row[2], row[3], row[4], row[5]
        owner = row[6] if len(row) > 6 else ""
        affectation = row[7] if len(row) > 7 else ""
        status = row[8] if len(row) > 8 else ""
        key = str(numero) if numero else f"__no_numero_{id(row)}"
        if key not in groups:
            groups[key] = {
                "fournisseur": fournisseur, "demandeur": demandeur, "numero": numero, "date": date,
                "montant_ht": 0, "owner": owner or "", "affectation": affectation or "", "status": status or "",
            }
            order.append(key)
        g = groups[key]
        if isinstance(montant_ht, (int, float)):
            g["montant_ht"] += montant_ht
        if not g["owner"] and owner:
            g["owner"] = owner
        if not g["affectation"] and affectation:
            g["affectation"] = affectation
        if not g["status"] and status:
            g["status"] = status

    return [
        (g["fournisseur"], g["demandeur"], g["numero"], g["date"], g["montant_ht"], g["owner"], g["affectation"], g["status"])
        for g in (groups[k] for k in order)
    ]


def _fill_invoice_summary_sheet(ws: Worksheet, invoice_rows: list[tuple], source_folder: str) -> Worksheet:
    """One row per invoice (no repeated invoice numbers), with the invoice's total Montant HT."""
    header_row = _title_block(
        ws,
        "Résumé par facture (1 ligne / facture)",
        f"Dossier source : {source_folder}  |  Factures : {len(invoice_rows)}",
        span_cols=8,
    )
    headers = ["Fournisseur", "Demandeur", "Numéro facture", "Date de facture", "Montant HT", "Owner CPM / BA", "Affectation", "Status"]
    _write_header(ws, headers, row=header_row)
    _add_dropdown(ws, OWNER_OPTIONS, column_index=6, first_row=header_row + 1, title="Owner CPM / BA")
    _add_dropdown(ws, STATUS_OPTIONS, column_index=8, first_row=header_row + 1, title="Status")
    r = header_row + 1
    for row_vals in invoice_rows:
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if c_idx == 5 and isinstance(val, (int, float)):
                cell.number_format = NUMBER_FORMAT
        r += 1

    last_data_row = r - 1
    if last_data_row >= header_row + 1:
        _protect_extracted_columns(
            ws, locked_columns=[1, 2, 3, 4, 5], editable_columns=[6, 7, 8],
            first_row=header_row + 1, last_row=last_data_row,
        )

    if invoice_rows:
        total_row = r
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        total_ht = sum(row[4] for row in invoice_rows if isinstance(row[4], (int, float)))
        cell = ws.cell(row=total_row, column=5, value=total_ht)
        cell.number_format = NUMBER_FORMAT
        cell.font = TOTAL_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=c).fill = TOTAL_FILL
            ws.cell(row=total_row, column=c).border = THIN_BORDER
        r += 1

    _band_rows(ws, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws, start_row=header_row)
    return ws


def _safe_save(wb: Workbook, output_path: Path) -> None:
    try:
        wb.save(str(output_path))
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot write '{output_path.name}' — it looks like the file is currently open "
            f"in Excel (or another program). Close it and try again."
        ) from exc


def write_workbook(
    results: list[PdfResult],
    output_path: Path,
    source_folder: str,
    factures: list[Facture] | None = None,
) -> Path:
    factures = factures or []
    matched_names = {f.file_name for f in factures}
    generic_results = [r for r in results if r.file_name not in matched_names]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Never lose previously extracted invoices: read back what's already in
    # the file (if any) and only append rows for genuinely new invoice numbers.
    existing_rows, existing_numeros = _load_existing_factures_sheet(output_path)
    new_factures = [f for f in factures if not (f.numero and f.numero in existing_numeros)]
    skipped = len(factures) - len(new_factures)
    if skipped:
        logger.info("%d invoice(s) already present in %s — not re-added.", skipped, output_path.name)

    new_rows: list[tuple] = []
    for fac in new_factures:
        new_rows.extend(_facture_to_rows(fac))

    all_rows = existing_rows + new_rows
    invoice_count = len(existing_numeros | {f.numero for f in new_factures if f.numero})

    wb = Workbook()
    wb.remove(wb.active)  # we'll add sheets explicitly in the right order

    # ------------------------------------------------------- Factures / Lignes
    if all_rows:
        ws_fac = wb.create_sheet("Factures")
        _fill_factures_sheet(ws_fac, all_rows, source_folder, invoice_count)
        wb.active = 0

        ws_inv = wb.create_sheet("Résumé Factures")
        _fill_invoice_summary_sheet(ws_inv, _group_rows_by_invoice(all_rows), source_folder)

    # If nothing was auto-recognized, fall back fully to the generic report
    if not generic_results:
        if not all_rows:
            wb.create_sheet("Summary")  # keep a sheet so the workbook isn't empty
        _safe_save(wb, output_path)
        return output_path

    # ---------------------------------------------------------------- Summary
    ws_summary = wb.create_sheet("Summary")
    header_row = _title_block(
        ws_summary,
        "PDF Extraction Summary",
        f"Source folder: {source_folder}  |  Files (not matching a known template): {len(generic_results)}  |  "
        f"Generated: {generic_results[0].extracted_at if generic_results else ''}",
        span_cols=7,
    )
    headers = ["File name", "Pages", "Tables found", "Fields found", "Dates found", "Amounts found", "Status"]
    _write_header(ws_summary, headers, row=header_row)
    r = header_row + 1
    for res in generic_results:
        status = "Error" if res.error else "OK"
        row_vals = [
            res.file_name,
            res.page_count,
            len(res.tables),
            len(res.key_values),
            len(res.dates_found),
            len(res.amounts_found),
            res.error if res.error else status,
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_summary.cell(row=r, column=c_idx, value=val)
            if res.error:
                cell.fill = ERROR_FILL
        r += 1
    _band_rows(ws_summary, header_row + 1, r - 1, len(headers))
    _autosize(ws_summary, start_row=header_row)

    # ----------------------------------------------------------------- Fields
    ws_fields = wb.create_sheet("Fields")
    header_row = _title_block(
        ws_fields, "Extracted Key / Value Fields", "One row per field found in each document.", span_cols=4
    )
    headers = ["File name", "Page count", "Field", "Value"]
    _write_header(ws_fields, headers, row=header_row)
    r = header_row + 1
    for res in generic_results:
        for key, value in res.key_values.items():
            ws_fields.cell(row=r, column=1, value=res.file_name)
            ws_fields.cell(row=r, column=2, value=res.page_count)
            ws_fields.cell(row=r, column=3, value=key)
            ws_fields.cell(row=r, column=4, value=value)
            r += 1
    _band_rows(ws_fields, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws_fields, start_row=header_row)

    # ----------------------------------------------------------------- Tables
    ws_tables = wb.create_sheet("Tables")
    header_row = _title_block(
        ws_tables, "Extracted Tables (combined)", "All tables from all documents, stacked together.", span_cols=6
    )
    headers = ["File name", "Page", "Table #", "Row #", "Column", "Value"]
    _write_header(ws_tables, headers, row=header_row)
    r = header_row + 1
    for res in generic_results:
        for tbl in res.tables:
            for row_idx, row in enumerate(tbl.rows, start=1):
                for col_idx, val in enumerate(row):
                    col_name = tbl.headers[col_idx] if col_idx < len(tbl.headers) else f"col_{col_idx+1}"
                    ws_tables.cell(row=r, column=1, value=res.file_name)
                    ws_tables.cell(row=r, column=2, value=tbl.page_number)
                    ws_tables.cell(row=r, column=3, value=tbl.table_index)
                    ws_tables.cell(row=r, column=4, value=row_idx)
                    ws_tables.cell(row=r, column=5, value=col_name)
                    ws_tables.cell(row=r, column=6, value=val)
                    r += 1
    _band_rows(ws_tables, header_row + 1, max(r - 1, header_row + 1), len(headers))
    _autosize(ws_tables, start_row=header_row)

    # ------------------------------------------------- One clean sheet / table
    used_names = {"Factures", "Résumé Factures", "Summary", "Fields", "Tables"}
    for res in generic_results:
        for tbl in res.tables:
            base = f"{Path(res.file_name).stem}_p{tbl.page_number}_t{tbl.table_index}"
            sheet_name = base[:31]
            suffix = 1
            while sheet_name in used_names:
                sheet_name = f"{base[:28]}_{suffix}"[:31]
                suffix += 1
            used_names.add(sheet_name)

            ws_t = wb.create_sheet(sheet_name)
            header_row_t = _title_block(
                ws_t, f"{res.file_name}", f"Page {tbl.page_number} — Table {tbl.table_index}", span_cols=max(1, len(tbl.headers))
            )
            _write_header(ws_t, tbl.headers, row=header_row_t)
            rr = header_row_t + 1
            for row in tbl.rows:
                for c_idx, val in enumerate(row, start=1):
                    ws_t.cell(row=rr, column=c_idx, value=val)
                rr += 1
            _band_rows(ws_t, header_row_t + 1, max(rr - 1, header_row_t + 1), max(1, len(tbl.headers)))
            _autosize(ws_t, start_row=header_row_t)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _safe_save(wb, output_path)
    return output_path
